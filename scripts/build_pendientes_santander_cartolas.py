#!/usr/bin/env python3
"""
Arma filas listas para la hoja Pendientes (ID, Fecha, Comercio, Monto, Tarjeta, Banco, Email_ID, Procesado)
a partir de:
  - Cartola cuenta corriente consolidada (.xlsx)
  - Cartolas cuenta vista consolidadas (.pdf)

Uso:
  PYTHONPATH=.vendor python3 scripts/build_pendientes_santander_cartolas.py

Salida:
  imports/pendientes_santander_cartolas_oct2025_mar2026.csv

Ajusta las rutas DEFAULT_* si cambian los archivos en Descargas.
"""

from __future__ import annotations

import csv
import hashlib
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
VENDOR = ROOT / ".vendor"
if VENDOR.is_dir():
    sys.path.insert(0, str(VENDOR))

from openpyxl import load_workbook  # noqa: E402

try:
    from pypdf import PdfReader  # noqa: E402
except ImportError as e:  # pragma: no cover
    raise SystemExit("Instala dependencias: pip install --target=.vendor openpyxl pypdf") from e

# Rutas por defecto (Descargas del usuario)
DEFAULT_XLSX = Path.home() / "Downloads" / "Cartola_Consolidada_Oct2025_Mar2026.xlsx"
DEFAULT_PDF = Path.home() / "Downloads" / "Cartolas_Consolidadas_70-75.pdf"

MES_ES = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}


def stable_id(parts: list[str]) -> str:
    h = hashlib.sha1("|".join(parts).encode("utf-8")).hexdigest()[:12]
    return f"bf-san-{h}"


def parse_mes_celda(s: str) -> tuple[int, int]:
    parts = (s or "").strip().split()
    if len(parts) < 2:
        raise ValueError(f"MES inválido: {s!r}")
    mo = MES_ES.get(parts[0].lower().strip())
    if not mo:
        raise ValueError(f"MES desconocido: {s!r}")
    y = int(parts[1])
    return y, mo


def fecha_iso_from_ddmm(ddmm: str, year_hint: int, month_hint: int) -> str:
    d, m = map(int, ddmm.split("/"))
    # Si el movimiento cae en otro mes calendario respecto al encabezado de cartola,
    # igualmente el par (d,m) + año del encabezado suele ser correcto para cartolas mensuales.
    y = year_hint
    try:
        return date(y, m, d).isoformat()
    except ValueError:
        # fallback: año siguiente/previo raramente necesario
        for dy in (-1, 1):
            try:
                return date(y + dy, m, d).isoformat()
            except ValueError:
                continue
    return date(year_hint, month_hint, min(d, 28)).isoformat()


def parse_excel_amount(cell) -> float | None:
    if cell is None or cell == "":
        return None
    if isinstance(cell, (int, float)):
        return float(cell)
    s = str(cell).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


@dataclass
class RowOut:
    id: str
    fecha: str
    comercio: str
    monto: float
    tarjeta: str
    banco: str
    email_id: str
    procesado: str = "NO"


def read_corriente(xlsx: Path) -> list[RowOut]:
    wb = load_workbook(xlsx, data_only=True)
    ws = wb["Movimientos Consolidados"]
    seen: set[tuple] = set()
    out: list[RowOut] = []

    for r in range(2, ws.max_row + 1):
        mes = ws.cell(r, 1).value
        fd = ws.cell(r, 2).value
        suc = ws.cell(r, 3).value or ""
        desc = (ws.cell(r, 4).value or "").strip()
        ndoc = ws.cell(r, 5).value
        cargo = parse_excel_amount(ws.cell(r, 6).value)
        abono = parse_excel_amount(ws.cell(r, 7).value)

        if not desc and cargo is None and abono is None:
            continue
        if not fd:
            continue

        y_hint, m_hint = parse_mes_celda(str(mes))
        ddmm = str(fd).strip()
        if not re.match(r"^\d{1,2}/\d{1,2}$", ddmm):
            continue

        fecha_iso = fecha_iso_from_ddmm(ddmm, y_hint, m_hint)

        if cargo is not None and cargo > 0 and abono is not None and abono > 0:
            # raro en cartola; conservar ambos como filas separadas con sufijo
            legs = [("cargo", cargo), ("abono", abono)]
        elif cargo is not None and cargo > 0:
            legs = [("cargo", cargo)]
        elif abono is not None and abono > 0:
            legs = [("abono", abono)]
        else:
            continue

        doc_txt = f" doc:{ndoc}" if ndoc not in (None, "") else ""
        suc_txt = f" suc:{suc}" if str(suc).strip() else ""

        for kind, amt in legs:
            dedupe_key = (fecha_iso, desc, kind, round(amt, 2), str(ndoc or ""))
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)

            tag = "Cargo" if kind == "cargo" else "Abono"
            comercio = f"[CC {tag}] {desc}{doc_txt}{suc_txt}".strip()
            rid = stable_id(["cte", fecha_iso, desc, kind, str(amt), str(ndoc or ""), suc_txt])
            out.append(
                RowOut(
                    id=rid,
                    fecha=fecha_iso,
                    comercio=comercio,
                    monto=float(amt),
                    tarjeta="Cuenta corriente",
                    banco="Santander",
                    email_id="",
                )
            )
    return out


def parse_clp_amount_token(tok: str) -> float:
    t = tok.replace(".", "").replace(",", ".")
    return float(t)


def resolve_date_in_range(d: int, m: int, desde: date, hasta: date) -> date | None:
    for y in range(desde.year - 1, hasta.year + 2):
        try:
            cand = date(y, m, d)
        except ValueError:
            continue
        if desde <= cand <= hasta:
            return cand
    return None


def read_vista_pdf(pdf: Path) -> list[RowOut]:
    reader = PdfReader(str(pdf))
    text = "\n".join((p.extract_text() or "") for p in reader.pages)
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    out: list[RowOut] = []
    seen: set[tuple] = set()

    desde: date | None = None
    hasta: date | None = None
    in_mov = False
    last_dm: tuple[int, int] | None = None

    for ln2 in lines:
        msec = re.match(
            r"^(\d+)\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\b",
            ln2,
        )
        if msec:
            _, ds, hs = msec.groups()
            desde = datetime.strptime(ds, "%d/%m/%Y").date()
            hasta = datetime.strptime(hs, "%d/%m/%Y").date()
            in_mov = False
            last_dm = None
            continue

        if ln2.startswith("MOVIMIENTO DE SU CUENTA"):
            in_mov = True
            continue

        if ln2.startswith("Resumen de Comisiones"):
            in_mov = False
            continue

        if not in_mov or desde is None or hasta is None:
            continue

        if ln2.startswith("--- Saldo Dia ---"):
            continue

        desc_part, amount, dm = parse_vista_movement_line(ln2, last_dm)
        if desc_part and amount is not None:
            dday, dmonth = dm
            resolved = resolve_date_in_range(dday, dmonth, desde, hasta)
            if not resolved:
                resolved = hasta
            fecha_iso = resolved.isoformat()
            comercio = f"[Vista] {desc_part}".strip()
            dedupe_key = (fecha_iso, comercio, round(amount, 2))
            if dedupe_key not in seen:
                seen.add(dedupe_key)
                rid = stable_id(["vista", fecha_iso, comercio, str(amount)])
                out.append(
                    RowOut(
                        id=rid,
                        fecha=fecha_iso,
                        comercio=comercio,
                        monto=float(amount),
                        tarjeta="Cuenta vista",
                        banco="Santander",
                        email_id="",
                    )
                )
            last_dm = dm

    return out


def parse_vista_movement_line(line: str, last_dm: tuple[int, int] | None) -> tuple[str, float | None, tuple[int, int]]:
    """
    Devuelve (descripcion_sin_monto, monto, (dia, mes)).
    El PDF suele pegar monto+fecha: 35.31801/10
    """
    if "MENSAJES" in line or "INFORMESE" in line:
        return "", None, last_dm or (1, 1)

    m_end = re.search(r"(\d{2}/\d{2})$", line)
    if m_end:
        dm_s = m_end.group(1)
        left = line[: m_end.start()].rstrip()
        dday, dmonth = map(int, dm_s.split("/"))
        mam = re.search(r"(\d{1,3}(?:\.\d{3})+|\d+)\s*$", left)
        if not mam:
            return "", None, (dday, dmonth)
        amt = parse_clp_amount_token(mam.group(1))
        desc = left[: mam.start()].strip()
        return desc, amt, (dday, dmonth)

    if last_dm is None:
        return "", None, (1, 1)

    mam = re.search(r"(\d{1,3}(?:\.\d{3})+|\d+)\s*$", line)
    if not mam:
        return "", None, last_dm
    amt = parse_clp_amount_token(mam.group(1))
    desc = line[: mam.start()].strip()
    return desc, amt, last_dm


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    pdf = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_PDF

    if not xlsx.is_file():
        raise SystemExit(f"No existe Excel: {xlsx}")
    if not pdf.is_file():
        raise SystemExit(f"No existe PDF: {pdf}")

    rows = read_corriente(xlsx) + read_vista_pdf(pdf)
    rows.sort(key=lambda r: (r.fecha, r.id))

    out_dir = ROOT / "imports"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / "pendientes_santander_cartolas_oct2025_mar2026.csv"

    with out_path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["ID", "Fecha", "Comercio", "Monto", "Tarjeta", "Banco", "Email_ID", "Procesado"])
        for r in rows:
            w.writerow([r.id, r.fecha, r.comercio, r.monto, r.tarjeta, r.banco, r.email_id, r.procesado])

    print(f"OK {len(rows)} filas -> {out_path}")


if __name__ == "__main__":
    main()
